"""
RFID Stock Reconciliation Service.
Replaces the "RFID Final" n8n workflow.

Flow: Telegram trigger (XLS file) → parse → normalize barcodes →
      query Odoo CTL/MUON stock → compare → generate XLSX report →
      send back via Telegram.
"""

import io
import re
import logging
import openpyxl
from services.odoo_client import odoo

logger = logging.getLogger(__name__)

# Barcodes to exclude (ported from "Lọc 3code đã mất lâu" node)
EXCLUDED_BARCODES = {
    "8968900015796",
    "8968900009665",
    "8968900015628",
    "8968900008446",
    "8968900011279",
    "8968900009030",
    "8968900009764",
    "8968900008705",
    "8968900012566",
    "8968900010920",
    "8968900016618",
}


def _normalize_epc_to_barcode(ecp: str) -> str:
    """
    Normalize EPC code to barcode by removing up to 3 leading zeros.
    Ported from: RFID Final / Normalize ECP -> Barcode node.
    """
    raw = str(ecp or "").strip()
    if not raw:
        return ""
    return re.sub(r"^0{1,3}", "", raw)


def _parse_excel_barcodes(file_bytes: bytes) -> list:
    """
    Parse an XLS/XLSX file to extract barcode data from Sheet1.
    Ported from: RFID Final / Extract from File node.
    Supports both .xls (binary/xlrd) and .xlsx (zip/openpyxl) formats.
    """
    rows = []

    # Try openpyxl first (.xlsx format)
    try:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True)
        ws = wb.active or wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        # Fall back to xlrd for .xls (binary) format
        try:
            import xlrd
            xls_wb = xlrd.open_workbook(file_contents=file_bytes)
            xls_ws = xls_wb.sheet_by_index(0)
            rows = [xls_ws.row_values(r) for r in range(xls_ws.nrows)]
            logger.info(f"Parsed .xls file with xlrd: {xls_ws.nrows} rows")
        except Exception as e:
            logger.error(f"Failed to parse Excel file with both openpyxl and xlrd: {e}")
            raise ValueError(f"Cannot parse Excel file: {e}")

    if not rows:
        return []

    # Get header row, find ECP column
    headers = [str(h or "").strip() for h in rows[0]]
    ecp_col = None
    for i, h in enumerate(headers):
        if h.upper() in ("ECP", "EPC", "BARCODE"):
            ecp_col = i
            break

    # If no matching header, use the first column
    if ecp_col is None:
        ecp_col = 0

    barcodes = []
    for row in rows[1:]:  # skip header
        if ecp_col < len(row):
            raw = str(row[ecp_col] or "").strip()
            bc = _normalize_epc_to_barcode(raw)
            if bc:
                barcodes.append(bc)

    return barcodes


def reconcile(file_bytes: bytes) -> dict:
    """
    Perform RFID stock reconciliation.
    Ported from: RFID Final workflow (complete flow).

    Args:
        file_bytes: Raw bytes of the uploaded XLS/XLSX file.

    Returns:
        dict with summary_text and xlsx_bytes for the report.
    """
    # ─── 1) Parse and normalize barcodes ───
    raw_barcodes = _parse_excel_barcodes(file_bytes)
    scanned_set = set(raw_barcodes)  # deduplicate
    total_scanned = len(scanned_set)

    logger.info(f"RFID scan: {total_scanned} unique barcodes parsed")

    # ─── 2) Query Odoo: CTL/Stock (location_id = 63) ───
    ctl_records = odoo.search_read(
        "stock.quant",
        [
            ("x_studio_barcode", "!=", False),
            ("location_id.id", "=", 63),
        ],
        fields=["id", "display_name", "x_studio_barcode", "product_id"],
        limit=2000,
    )
    total_ctl = len(ctl_records)
    logger.info(f"CTL/Stock records: {total_ctl}")

    # ─── 3) Query Odoo: MUON/Stock ───
    # NOTE: display_name on stock.quant includes the product (e.g. "[MUON/Stock] Product A").
    # Use location_id.complete_name to filter by location path reliably.
    muon_records = odoo.search_read(
        "stock.quant",
        [
            ("x_studio_barcode", "!=", False),
            ("location_id.complete_name", "=", "MUON/Stock"),
        ],
        fields=["id", "display_name", "x_studio_barcode", "product_id"],
        limit=2000,
    )
    total_muon = len(muon_records)
    logger.info(f"MUON/Stock records: {total_muon}")

    # ─── 4) Index CTL and MUON ───
    ctl_by_bc = {}
    for rec in ctl_records:
        bc = str(rec.get("x_studio_barcode") or "").strip()
        if bc and bc not in EXCLUDED_BARCODES:
            ctl_by_bc[bc] = rec

    muon_barcodes = {
        str(rec.get("x_studio_barcode") or "").strip()
        for rec in muon_records
        if str(rec.get("x_studio_barcode") or "").strip()
    }

    # ─── 5) Find missing items ───
    missing_items = []
    for bc, rec in ctl_by_bc.items():
        if bc not in muon_barcodes and bc not in scanned_set:
            prod = rec.get("product_id")
            pid = prod[0] if isinstance(prod, (list, tuple)) and prod else None
            pname = prod[1] if isinstance(prod, (list, tuple)) and len(prod) > 1 else None
            missing_items.append((bc, pid, pname))

    missing_count = len(missing_items)

    # ─── 6) Generate summary text ───
    summary = (
        "📦 Stock scan summary\n\n"
        f"• Total scanned barcodes (unique): {total_scanned}\n"
        f"• Total CTL/Stock items: {total_ctl}\n"
        f"• Total MUON/Stock items: {total_muon}\n"
        f"• Missing (in CTL, not in MUON, not scanned): {missing_count}\n"
    )

    # Preview of top 20 missing
    if missing_items:
        summary += "\nTop missing items:\n"
        for bc, pid, pname in missing_items[:20]:
            pid_str = f"[{pid}]" if pid else "[?]"
            summary += f"- {pid_str} {pname or ''}\n"
        summary += f"\n(Full list attached as XLSX — {missing_count} items)"

    # ─── 7) Generate XLSX report ───
    xlsx_bytes = _generate_xlsx_report(missing_items)

    return {
        "summary_text": summary,
        "xlsx_bytes": xlsx_bytes,
        "missing_count": missing_count,
        "total_scanned": total_scanned,
        "total_ctl": total_ctl,
        "total_muon": total_muon,
    }


def _generate_xlsx_report(missing_items: list) -> bytes:
    """
    Generate an XLSX report of missing stock items.
    Ported from: RFID Final / Code in Python (Beta)1 node.
    Uses openpyxl instead of manual zipfile construction.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Missing Stock"

    # Header
    ws.append(["Barcode", "Product ID", "Product Name"])

    # Style header
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    # Data rows
    for bc, pid, pname in missing_items:
        ws.append([bc, pid, pname or ""])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or "")) for cell in col)
        col_letter = col[0].column_letter
        ws.column_dimensions[col_letter].width = min(max_len + 2, 50)

    # Save to bytes
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
